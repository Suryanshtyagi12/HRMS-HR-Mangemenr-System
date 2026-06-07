import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(report_type: str, month: int, year: int, data: dict, ai_summary: dict) -> bytes:
    wb = openpyxl.Workbook()
    
    # ------------------
    # Sheet 1: Summary
    # ------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Styles
    title_font = Font(name='Arial', size=18, bold=True, color='4F46E5')
    header_font = Font(name='Arial', size=14, bold=True, color='1E293B')
    bold_font = Font(bold=True)
    
    ws_summary['A1'] = f"HRMS Pro - {report_type.title()} Report"
    ws_summary['A1'].font = title_font
    
    ws_summary['A3'] = "AI Insights & Executive Summary"
    ws_summary['A3'].font = header_font
    
    row = 5
    ws_summary[f'A{row}'] = "Overall Health"
    ws_summary[f'A{row}'].font = bold_font
    ws_summary[f'B{row}'] = ai_summary.get('overall_health', 'N/A')
    row += 2
    
    ws_summary[f'A{row}'] = "Executive Summary"
    ws_summary[f'A{row}'].font = bold_font
    ws_summary[f'B{row}'] = ai_summary.get('executive_summary', '')
    ws_summary[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws_summary.row_dimensions[row].height = 40
    row += 2
    
    ws_summary[f'A{row}'] = "Highlights"
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    for h in ai_summary.get('highlights', []):
        ws_summary[f'B{row}'] = f"- {h}"
        row += 1
    row += 1
    
    ws_summary[f'A{row}'] = "Concerns"
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    for c in ai_summary.get('concerns', []):
        ws_summary[f'B{row}'] = f"- {c}"
        row += 1
    row += 1
    
    ws_summary[f'A{row}'] = "Recommendations"
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    for r in ai_summary.get('recommendations', []):
        ws_summary[f'B{row}'] = f"- {r}"
        row += 1
        
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 80
    
    # ------------------
    # Sheet 2: Data
    # ------------------
    ws_data = wb.create_sheet(title="Data")
    
    raw_rows = data.get("raw_rows", [])
    if raw_rows:
        headers = list(raw_rows[0].keys())
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
            
        # Write rows
        for row_idx, row_data in enumerate(raw_rows, 2):
            for col_idx, header in enumerate(headers, 1):
                ws_data.cell(row=row_idx, column=col_idx, value=str(row_data.get(header, "")))
                
        # Auto column widths
        for col_idx, header in enumerate(headers, 1):
            letter = get_column_letter(col_idx)
            max_len = max([len(str(r.get(header, ""))) for r in raw_rows] + [len(header)])
            ws_data.column_dimensions[letter].width = min(max_len + 2, 50) # Cap at 50
            
        # Auto filter & freeze
        ws_data.auto_filter.ref = ws_data.dimensions
        ws_data.freeze_panes = "A2"
    else:
        ws_data['A1'] = "No data available"

    # ------------------
    # Sheet 3: Chart
    # ------------------
    ws_chart = wb.create_sheet(title="Chart")
    ws_chart['A1'] = "Data Visualization"
    ws_chart['A1'].font = title_font
    ws_chart['A3'] = "See the frontend dashboard for rich interactive charts."
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
